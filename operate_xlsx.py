"""
此文件用于xlsx文件的处理
"""
import openpyxl  # 导入openpyxl模块
import numpy as np

# 创建workbook对象
path = r"C:/Users/28101/Desktop/test/test_1.xlsx"
wb = openpyxl.load_workbook(path)
# 得到当前活跃表单的对象（打开该xlsx文件，直接看到的表单就为活跃表单）
ws = wb.active
# 获取最大的行
r_max = ws.max_row
row_num = 17045
# row_num = 10
print("r_max:{}".format(r_max))
print(type(ws.cell(row=row_num, column=1).value))
index_all = []
for i in range(row_num):
    index_all.append(ws.cell(row=i+1, column=1).value)
print(index_all)

num = 2
for i in range(r_max):
    index_row = 0
    for j in range(len(index_all)):
        if ws.cell(row=i+1, column=6).value == index_all[j]:
            index_row = j + 1
            if i == 0:
                ws.cell(row=index_row, column=num).value = str(ws.cell(row=i + 1, column=7).value)
            else:
                if ws.cell(row=i+1, column=6).value == ws.cell(row=i, column=6).value:
                    num += 1
                else:
                    num = 2
                ws.cell(row=index_row, column=num).value = str(ws.cell(row=i+1, column=7).value)
            # wb.save(r"C:/Users/28101/Desktop/test/test_1.xlsx")
    print(ws.cell(row=i+1, column=6).value)
wb.save(path)
print("完成")





# from openpyxl import load_workbook
#
# wb = load_workbook(r"C:\Users\28101\Desktop\test\demo.xlsx")
# sheet = wb.active
#
# sheet['A1'] = 'Devansh Sharma'
# sheet['A2'] = 'hello world'
#
# sheet.cell(row=2, column=2).value = 5
#
# wb.save(r"C:\Users\28101\Desktop\test\demo.xlsx")
#
# print("运行结束！")